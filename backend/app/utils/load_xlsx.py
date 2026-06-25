from langchain_core.documents import Document

def load_xlsx(filepath: str) -> list[Document]:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    documents = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_data = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                row_str = ", ".join([str(cell) if cell is not None else "" for cell in row])
                sheet_data.append(row_str)
        content = "\n".join(sheet_data)
        if content.strip():
            documents.append(Document(page_content=content, metadata={"source": filepath, "sheet": sheet_name}))
    return documents